from __future__ import annotations

import csv
import io
import os
import re
from datetime import date, datetime
from typing import List, Optional, Tuple

from fastapi import UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from app.services.llm.llm_client import (
    ImagePositionItem,
    LLMProvider,
    PositionItem,
    extract_positions,
    extract_positions_from_image as llm_extract_positions_from_image,
)

CELL_REF_PATTERN = re.compile(r"^([A-Za-z]+)(\d+)?$")
SUPPORTED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}


class CropSettings(BaseModel):
    sheet: Optional[str] = None
    top_left: str = Field(..., min_length=1)
    bottom_right: str = Field(..., min_length=1)


def _column_to_index(column: str) -> int:
    result = 0
    for char in column.upper():
        if not ("A" <= char <= "Z"):
            raise ValueError("Invalid column reference")
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def _parse_cell_ref(value: str, require_row: bool) -> Tuple[int, Optional[int]]:
    match = CELL_REF_PATTERN.match(value.strip())
    if not match:
        raise ValueError("Invalid cell reference")
    column = _column_to_index(match.group(1))
    row_str = match.group(2)
    if require_row and not row_str:
        raise ValueError("Row is required for this cell reference")
    row = int(row_str) if row_str else None
    return column, row


def _normalize_cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _row_has_value(row: List[str], start_col: int, end_col: int) -> bool:
    for idx in range(start_col, end_col + 1):
        if idx < len(row) and row[idx].strip():
            return True
    return False


def _trim_trailing_empty_rows(rows: List[List[str]]) -> List[List[str]]:
    while rows and all(not cell.strip() for cell in rows[-1]):
        rows.pop()
    return rows


def _crop_rows(
    rows: List[List[str]],
    top_left: str,
    bottom_right: str,
) -> List[List[str]]:
    start_col, start_row = _parse_cell_ref(top_left, require_row=True)
    end_col, end_row = _parse_cell_ref(bottom_right, require_row=False)

    if end_col < start_col:
        raise ValueError("Bottom-right column must be after top-left column")

    start_row_index = start_row - 1
    start_col_index = start_col - 1
    end_col_index = end_col - 1

    if start_row_index >= len(rows):
        return []

    if end_row is None:
        last_row_index = -1
        for idx in range(start_row_index, len(rows)):
            if _row_has_value(rows[idx], start_col_index, end_col_index):
                last_row_index = idx
        if last_row_index < start_row_index:
            return []
    else:
        last_row_index = min(end_row - 1, len(rows) - 1)

    cropped: List[List[str]] = []
    for idx in range(start_row_index, last_row_index + 1):
        row = rows[idx]
        cropped_row = []
        for col_idx in range(start_col_index, end_col_index + 1):
            cropped_row.append(row[col_idx] if col_idx < len(row) else "")
        cropped.append(cropped_row)

    return _trim_trailing_empty_rows(cropped)


def _read_csv_bytes(data: bytes) -> List[List[str]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = data.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    return [[cell.strip() for cell in row] for row in reader]


def _select_worksheet(workbook, sheet: Optional[str]):
    if not sheet or sheet.lower() == "first":
        return workbook.worksheets[0]
    if sheet.isdigit():
        index = int(sheet) - 1
        if index < 0 or index >= len(workbook.worksheets):
            raise ValueError("Sheet index out of range")
        return workbook.worksheets[index]
    if sheet in workbook.sheetnames:
        return workbook[sheet]
    raise ValueError("Sheet not found")


def _read_xlsx_bytes(data: bytes, sheet: Optional[str]) -> List[List[str]]:
    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    worksheet = _select_worksheet(workbook, sheet)

    rows: List[List[str]] = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append([_normalize_cell_value(cell) for cell in row])
    return rows


def is_image_file(filename: Optional[str], content_type: Optional[str] = None) -> bool:
    if content_type and content_type.startswith("image/"):
        return True
    if not filename:
        return False
    extension = os.path.splitext(filename)[1].lower()
    return extension in SUPPORTED_IMAGE_EXTENSIONS


def _infer_image_mime_type(filename: Optional[str]) -> str:
    if not filename:
        return "image/png"
    extension = os.path.splitext(filename)[1].lower()
    if extension in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if extension == ".webp":
        return "image/webp"
    return "image/png"


async def load_image_bytes(file: UploadFile) -> tuple[bytes, str]:
    data = await file.read()
    mime_type = ""
    if file.content_type and file.content_type.startswith("image/"):
        mime_type = file.content_type
    if not mime_type:
        mime_type = _infer_image_mime_type(file.filename)
    return data, mime_type


async def load_cropped_rows(
    file: UploadFile,
    crop: CropSettings,
) -> List[List[str]]:
    filename = (file.filename or "").lower()
    data = await file.read()

    if filename.endswith(".csv"):
        rows = _read_csv_bytes(data)
    elif filename.endswith(".xlsx"):
        rows = _read_xlsx_bytes(data, crop.sheet)
    elif filename.endswith(".xls"):
        raise ValueError("Unsupported file type .xls. Please export as .xlsx or .csv")
    else:
        raise ValueError("Unsupported file type. Please upload .csv or .xlsx")

    return _crop_rows(rows, crop.top_left, crop.bottom_right)


def _build_providers(provider_settings: List[dict]) -> List[LLMProvider]:
    return [
        LLMProvider(
            name=item.get("name", "provider"),
            base_url=item.get("base_url", ""),
            api_key=item.get("api_key", ""),
            model=item.get("model", ""),
        )
        for item in provider_settings
    ]


def merge_image_positions(
    positions: List[ImagePositionItem],
) -> tuple[List[ImagePositionItem], List[str]]:
    normalized: dict[str, ImagePositionItem] = {}
    conflicted: set[str] = set()

    for item in positions:
        ticker = item.ticker.strip().upper()
        if not ticker:
            continue
        try:
            average_cost = float(item.average_cost)
        except (TypeError, ValueError):
            continue
        if average_cost <= 0:
            continue
        quantity: Optional[float] = None
        if item.quantity is not None:
            try:
                quantity = float(item.quantity)
            except (TypeError, ValueError):
                quantity = None
            if quantity is not None and quantity <= 0:
                quantity = None

        average_cost = round(average_cost, 4)
        if quantity is not None:
            quantity = round(quantity, 4)

        existing = normalized.get(ticker)
        if existing:
            if existing.average_cost != average_cost:
                conflicted.add(ticker)
                continue
            if existing.quantity is not None and quantity is not None and existing.quantity != quantity:
                conflicted.add(ticker)
                continue
            if existing.quantity is None and quantity is not None:
                existing.quantity = quantity
            continue

        normalized[ticker] = ImagePositionItem(
            ticker=ticker,
            average_cost=average_cost,
            quantity=quantity,
        )

    for ticker in conflicted:
        normalized.pop(ticker, None)

    result = list(normalized.values())
    result.sort(key=lambda item: item.ticker)
    return result, sorted(conflicted)


async def extract_positions_from_rows(
    rows: List[List[str]],
    provider_settings: List[dict],
    timeout_seconds: int,
) -> List[PositionItem]:
    providers = _build_providers(provider_settings)

    positions = await extract_positions(
        rows,
        providers,
        timeout_seconds,
        caller="portfolio_import",
    )
    return [
        PositionItem(
            ticker=position.ticker.strip().upper(),
            quantity=position.quantity,
        )
        for position in positions
    ]


async def extract_positions_from_image(
    file: UploadFile,
    provider_settings: List[dict],
    timeout_seconds: int,
) -> List[ImagePositionItem]:
    providers = _build_providers(provider_settings)
    image_bytes, mime_type = await load_image_bytes(file)
    return await llm_extract_positions_from_image(
        image_bytes,
        mime_type,
        providers,
        timeout_seconds,
        caller="portfolio_import",
    )
